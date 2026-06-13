"""XLSX to Markdown conversion using openpyxl."""

from __future__ import annotations

import html
import io
import logging
import re
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.config import settings
from app.models.response import JsonlEvent, SheetJson

logger = logging.getLogger(__name__)

# Dangerous URL schemes that must be blocked to prevent injection attacks
_DANGEROUS_URL_SCHEMES = re.compile(
    r"^(?:javascript|vbscript|data|file|blob):",
    re.IGNORECASE,
)


def _sanitize_url(url: str | None) -> str | None:
    """Sanitize a URL by blocking dangerous schemes.

    Args:
        url: The raw URL string to sanitize, or None.

    Returns:
        The sanitized URL, "#" if the URL uses a dangerous scheme, or None.
    """
    if not url:
        return None
    stripped = url.strip()
    if _DANGEROUS_URL_SCHEMES.match(stripped):
        logger.warning("Blocked dangerous URL scheme in XLSX: %s", stripped[:50])
        return "#"
    return url


def _escape_md_table(text: str | None) -> str | None:
    r"""Escape characters that have special meaning in Markdown tables.

    Escapes `|` and `\` to prevent table structure corruption.

    Args:
        text: Raw cell text, or None.

    Returns:
        Text with Markdown table special characters escaped, or None.
    """
    if text is None:
        return None
    if not text:
        return text
    text = text.replace("\\", "\\\\")
    text = text.replace("|", "\\|")
    return text


def _cell_to_string(cell) -> str:
    """Convert an openpyxl cell value to its string representation.

    Handles different cell types: text, number, date, boolean, formula, None.

    Args:
        cell: An openpyxl cell object.

    Returns:
        The string representation of the cell value.
    """
    value = cell.value

    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    # String value (including formula results that are strings)
    return str(value)


def _get_merged_cell_value(ws, row: int, col: int) -> str:
    """Get the value of a cell, checking merged ranges first.

    If the cell is part of a merged range, returns the value of the
    top-left cell of that range.

    Note: In read_only mode, merged_cells is not available, so we
    just return the cell value directly.

    Args:
        ws: The openpyxl worksheet.
        row: Row number (1-based).
        col: Column number (1-based).

    Returns:
        The string representation of the cell value.
    """
    # In read_only mode, merged_cells is not available
    # We skip merged cell handling for better performance
    try:
        if hasattr(ws, "merged_cells") and ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                if (row, col) in merged_range:
                    top_cell = ws.cell(
                        row=merged_range.min_row, column=merged_range.min_col
                    )
                    return _cell_to_string(top_cell)
    except AttributeError:
        # ReadOnlyWorksheet doesn't have merged_cells attribute
        pass
    return _cell_to_string(ws.cell(row=row, column=col))


def _build_sheet_md(ws) -> str:
    """Convert a single worksheet to a Markdown table.

    Args:
        ws: An openpyxl worksheet object.

    Returns:
        A Markdown string with a heading and table representation of the sheet.
    """
    if ws.max_row == 0 or ws.max_column == 0:
        return f"# {html.escape(ws.title)}\n\n_(empty sheet)_"

    # Collect all values into a 2D array
    rows = []
    for row_idx in range(1, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            value = _get_merged_cell_value(ws, row_idx, col_idx)
            row_values.append(value)
        rows.append(row_values)

    # Build Markdown table
    lines = []

    # Header row
    header = "| " + " | ".join(_escape_md_table(v) for v in rows[0]) + " |"
    lines.append(header)

    # Separator row
    separator = "| " + " | ".join("---" for _ in rows[0]) + " |"
    lines.append(separator)

    # Data rows
    for row in rows[1:]:
        line = "| " + " | ".join(_escape_md_table(v) for v in row) + " |"
        lines.append(line)

    return f"# {html.escape(ws.title)}\n\n" + "\n".join(lines)


def convert_xlsx_to_md(file_bytes: bytes) -> str:
    """Pure Python XLSX → MD conversion (openpyxl).

    Converts all worksheets to Markdown tables, handling:
    - Multiple sheets
    - Merged cells
    - Different cell types (text, number, date, boolean, formula)
    - Empty cells
    - Special Markdown character escaping

    Args:
        file_bytes: The raw bytes of the XLSX file.

    Returns:
        A Markdown string representing the entire workbook.

    Raises:
        InvalidFileException: If the file is not a valid XLSX.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except InvalidFileException as e:
        logger.error("Invalid XLSX file: %s", e)
        raise

    # Log openpyxl warnings
    if hasattr(wb, "properties"):
        logger.debug(
            "XLSX properties: title=%s, creator=%s",
            wb.properties.title,
            wb.properties.creator,
        )

    sheets_md = []
    for ws in wb.worksheets:
        try:
            sheet_md = _build_sheet_md(ws)
            sheets_md.append(sheet_md)
        except Exception as e:
            logger.warning("Error converting sheet '%s': %s", ws.title, e)
            sheets_md.append(
                f"# {html.escape(ws.title)}\n\n"
                f"_(error converting sheet: {html.escape(str(e))})_"
            )

    wb.close()

    return "\n\n---\n\n".join(sheets_md)


def _extract_xlsx_content(file_bytes: bytes) -> list[tuple[int, str, list[list[str]]]]:
    """Extract XLSX content as sheets (tables).

    Args:
        file_bytes: Raw XLSX file bytes.

    Returns:
        List of tuples (sheet_num, title, cell_values_list_of_lists).
    """
    logger = logging.getLogger(__name__)
    print(f"[DEBUG] Starting _extract_xlsx_content with {len(file_bytes)} bytes", flush=True)

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        print(f"[DEBUG] Loaded workbook with {len(wb.worksheets)} worksheets", flush=True)
    except InvalidFileException as e:
        logger.error("Invalid XLSX file: %s", e)
        raise

    results: list[tuple[int, str, list[list[str]]]] = []
    sheet_num = 0

    print("[DEBUG] Starting loop over worksheets", flush=True)
    for ws in wb.worksheets:
        print(f"[DEBUG] Processing sheet '{ws.title}'", flush=True)
        try:
            # Collect all values into a 2D array
            rows = []
            max_row = ws.max_row if hasattr(ws, 'max_row') else 0
            max_col = ws.max_column if hasattr(ws, 'max_column') else 0
            print(f"[DEBUG] Sheet '{ws.title}': max_row={max_row}, max_col={max_col}", flush=True)

            for row_idx in range(1, max_row + 1):
                print(f"[DEBUG] Processing row {row_idx}", flush=True)
                row_values = []
                max_col_for_row = ws.max_column if hasattr(ws, 'max_column') else 0
                for col_idx in range(1, max_col_for_row + 1):
                    try:
                        value = _get_merged_cell_value(ws, row_idx, col_idx)
                        print(f"[DEBUG] Cell ({row_idx},{col_idx}): {value}", flush=True)
                        row_values.append(value)
                    except Exception as e:
                        print(f"[ERROR] Error reading cell ({row_idx},{col_idx}): {e}", flush=True)
                rows.append(row_values)

            try:
                sheet_num += 1
                print(f"[DEBUG] Sheet '{ws.title}' has {len(rows)} rows", flush=True)
                escaped_title = html.escape(ws.title)
                print(f"[DEBUG] Escaped title: {escaped_title}", flush=True)
                result_tuple = (sheet_num, escaped_title, rows)
                results.append(result_tuple)
                print(f"[DEBUG] Successfully appended! len(results)={len(results)}", flush=True)
            except Exception as e:
                import traceback
                print(f"[ERROR] Error saving sheet '{ws.title}': {e}", flush=True)
                traceback.print_exc()

        except Exception as e:
            print(f"[ERROR] Error in try block for sheet '{ws.title}': {e}", flush=True)
            import traceback
            traceback.print_exc()

    print(f"[DEBUG] Before close, len(results)={len(results)}", flush=True)
    for i, r in enumerate(results):
        print(f"[DEBUG] Result {i}: sheet_num={r[0]}, title='{r[1]}', rows={len(r[2])}", flush=True)
    wb.close()
    return results


def convert_xlsx_to_json(file_bytes: bytes) -> dict:
    """Convert XLSX to JSON format.

    Args:
        file_bytes: Raw XLSX file bytes.

    Returns:
        Dict with sheets and metadata in JSON structure.
    """

    results = _extract_xlsx_content(file_bytes)

    return {
        "format": "xlsx",
        "sheets": [
            SheetJson(
                name=sheet[1],
                data=[[_escape_md_table(str(cell_val)) if cell_val else "" for cell_val in row] for row in sheet[2]],
                headers=None
            ).model_dump()
            for sheet in results
        ],
        "metadata": {
            "format": "xlsx",
            "size_bytes": len(file_bytes),
            "sheet_count": len(results),
        },
    }


def convert_xlsx_to_jsonl(file_bytes: bytes) -> list[JsonlEvent]:
    """Convert XLSX to JSONL format with chunking.

    Args:
        file_bytes: Raw XLSX file bytes.

    Returns:
        List of JsonlEvent objects with start, chunk, and end events.
    """

    results = _extract_xlsx_content(file_bytes)

    return _to_jsonl(results)


def _to_jsonl(results: list[tuple[int, str, list[list[str]]]]) -> list[JsonlEvent]:
    """Convert extraction results to JSONL format with chunking.

    Args:
        results: List of (sheet_num, title, cell_values_list_of_lists) tuples.

    Returns:
        List of JsonlEvent objects with start, chunk, and end events.
    """

    # Convert to tabular text representation for chunking
    all_text: list[str] = []
    for sheet_num, title, rows in results:
        all_text.append(f"Sheet {sheet_num}: {title}")
        for row_idx, row_values in enumerate(rows, 1):
            row_str = "| " + " | ".join(str(v) if v else "" for v in row_values) + " |"
            all_text.append(f"Row {row_idx}:\n{row_str}")

    chunk_size = settings.jsonl_chunk_size

    # Start event
    events: list[JsonlEvent] = [JsonlEvent(
        type="start",
        metadata={"format": "xlsx"}
    )]

    if all_text:
        chunks: list[list[str]] = [all_text[i:i + chunk_size] for i in range(0, len(all_text), chunk_size)]

        for chunk in chunks:
            events.append(JsonlEvent(type="chunk", markdown_text="\n".join(chunk)))

    # End event
    events.append(JsonlEvent(
        type="end",
        metadata={"format": "xlsx", "total_sheets": len(results)}
    ))

    return events
