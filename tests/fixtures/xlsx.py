"""Fixtures for XLSX tests."""

__all__ = [
    "sample_xlsx_bytes",
    "sample_xlsx_simple_bytes",  # Nouveau fixture pour les tests Markdown simples
    "sample_xlsx_multi_sheet_bytes",
    "sample_xlsx_merged_cells_bytes",
    "sample_xlsx_dates_numbers_bytes",
    "sample_xlsx_special_chars_bytes",
    "sample_xlsx_empty_sheet_bytes",
]

import io
from datetime import date, datetime
import pytest
from openpyxl import Workbook


@pytest.fixture
def sample_xlsx_bytes() -> bytes:
    """Generate a minimal XLSX file in memory with one sheet and several cells."""
    wb: Workbook = Workbook()
    ws = wb.active  # type: ignore[misc]
    ws.title = "Feuille1"  # type: ignore[attr-defined]
    
    # En-têtes - plusieurs colonnes pour générer plus de texte
    headers = ["ID", "Nom", "Description_détaillée", "Prix_unitaire", "Quantité_disponible", "Catégorie_produit"]
    for col, header in enumerate(headers, 1):
        ws[f"A{col}"] = header  # type: ignore[index]
    
    # Données - beaucoup de lignes avec du texte long pour générer des chunks
    descriptions = [
        "Ceci est une description très détaillée pour le produit numéro {i}. Ce texte est suffisamment long pour dépasser la taille de chunk par défaut de 1024 caractères lors de la conversion JSONL.",
        "Description complète du produit avec beaucoup d'informations techniques et des spécifications détaillées pour tester correctement le chunking.",
        "Ceci est une description très longue et détaillée pour le produit numéro {i}. Le texte doit être assez long pour générer plusieurs chunks lors de la conversion en format JSONL.",
    ]
    
    for i in range(2, 51):  # Lignes 2 à 50 (49 lignes de données)
        ws[f"A{i}"] = i  # type: ignore[index]
        ws[f"B{i}"] = f"Produit {i}"  # type: ignore[index]
        desc_idx = (i - 2) % len(descriptions)
        ws[f"C{i}"] = descriptions[desc_idx].format(i=i)  # type: ignore[index]
        ws[f"D{i}"] = float(i * 1.5)  # type: ignore[index]
        ws[f"E{i}"] = i * 10  # type: ignore[index]
        ws[f"F{i}"] = f"Catégorie_{i % 5}"  # type: ignore[index]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@pytest.fixture
def sample_xlsx_simple_bytes() -> bytes:
    """Generate a minimal XLSX file with simple data for Markdown tests."""
    from openpyxl import Workbook

    wb: Workbook = Workbook()
    ws = wb.active
    ws.title = "Feuille1"
    
    # En-têtes simples pour les tests Markdown (ligne 1)
    ws["A1"] = "Nom"  # type: ignore[index]
    ws["B1"] = "Valeur"  # type: ignore[index]
    
    # Données simples (lignes 2 à 6)
    for i in range(1, 6):
        ws[f"A{i+1}"] = f"Produit {i}"  # type: ignore[index]
        ws[f"B{i+1}"] = i  # type: ignore[index]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@pytest.fixture
def sample_xlsx_multi_sheet_bytes() -> bytes:
    """Generate an XLSX file containing multiple sheets."""
    from openpyxl import Workbook

    wb: Workbook = Workbook()
    ws1 = wb.active  # type: ignore[misc]
    ws1.title = "Données"  # type: ignore[attr-defined]
    ws1["A1"] = "Produit"  # type: ignore[index]
    ws1["B1"] = "Prix"  # type: ignore[index]
    ws1["A2"] = "Pomme"  # type: ignore[index]
    ws1["B2"] = 1.5  # type: ignore[index]
    ws1["A3"] = "Orange"  # type: ignore[index]
    ws1["B3"] = 2.0  # type: ignore[index]

    ws2 = wb.create_sheet(title="Résumé")  # type: ignore[misc]
    ws2["A1"] = "Total"  # type: ignore[index]
    ws2["B1"] = 3.5  # type: ignore[index]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@pytest.fixture
def sample_xlsx_merged_cells_bytes() -> bytes:
    """Generate an XLSX file containing merged cells."""
    from openpyxl import Workbook

    wb: Workbook = Workbook()
    ws = wb.active  # type: ignore[misc]
    ws.title = "Fusionné"  # type: ignore[attr-defined]
    ws["A1"] = "En-tête fusionné"  # type: ignore[index]
    ws.merge_cells("A1:C1")  # type: ignore[attr-defined]
    ws["A2"] = "Col1"  # type: ignore[index]
    ws["B2"] = "Col2"  # type: ignore[index]
    ws["C2"] = "Col3"  # type: ignore[index]
    ws["A3"] = "v1"  # type: ignore[index]
    ws["B3"] = "v2"  # type: ignore[index]
    ws["C3"] = "v3"  # type: ignore[index]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@pytest.fixture
def sample_xlsx_dates_numbers_bytes() -> bytes:
    """Generate an XLSX file containing dates and numeric values."""
    from openpyxl import Workbook

    wb: Workbook = Workbook()
    ws = wb.active  # type: ignore[misc]
    ws.title = "Types"  # type: ignore[attr-defined]
    ws["A1"] = "Texte"  # type: ignore[index]
    ws["B1"] = "Nombre"  # type: ignore[index]
    ws["C1"] = "Date"  # type: ignore[index]
    ws["D1"] = "Booléen"  # type: ignore[index]
    ws["A2"] = "hello"  # type: ignore[index]
    ws["B2"] = 42.5  # type: ignore[index]
    ws["C2"] = date(2024, 1, 15)  # type: ignore[index]
    ws["D2"] = True  # type: ignore[index]
    ws["A3"] = "world"  # type: ignore[index]
    ws["B3"] = -10  # type: ignore[index]
    ws["C3"] = datetime(2024, 6, 30, 12, 0, 0)  # type: ignore[index]
    ws["D3"] = False  # type: ignore[index]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@pytest.fixture
def sample_xlsx_special_chars_bytes() -> bytes:
    """Generate an XLSX file containing Markdown special characters."""
    from openpyxl import Workbook

    wb: Workbook = Workbook()
    ws = wb.active  # type: ignore[misc]
    ws.title = "Spécial"  # type: ignore[attr-defined]
    ws["A1"] = "Colonne A"  # type: ignore[index]
    ws["B1"] = "Colonne B"  # type: ignore[index]
    ws["A2"] = "Texte avec | pipe"  # type: ignore[index]
    ws["B2"] = "Texte avec \\ backslash"  # type: ignore[index]
    ws["A3"] = "Ligne 1\nLigne 2"  # type: ignore[index]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@pytest.fixture
def sample_xlsx_empty_sheet_bytes() -> bytes:
    """Generate an XLSX file containing an empty sheet."""
    from openpyxl import Workbook

    wb: Workbook = Workbook()
    ws = wb.active  # type: ignore[misc]
    ws.title = "Vide"  # type: ignore[attr-defined]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
