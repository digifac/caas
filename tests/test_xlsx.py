"""Tests for XLSX to Markdown conversion."""

import io

import pytest
from app.converters.xlsx import (
    _escape_md_table,
    _sanitize_url,
    convert_xlsx_to_md,
)
from openpyxl import Workbook

# --- Tests _escape_md_table ---


class TestEscapeMdTable:
    def test_simple_text_unchanged(self):
        assert _escape_md_table("hello") == "hello"

    def test_pipe_escaped(self):
        assert _escape_md_table("a | b") == "a \\| b"

    def test_backslash_escaped(self):
        assert _escape_md_table("a \\ b") == "a \\\\ b"

    def test_pipe_and_backslash(self):
        assert _escape_md_table("\\|") == "\\\\\\|"

    def test_newlines_preserved(self):
        assert _escape_md_table("line1\nline2") == "line1\nline2"

    def test_empty_string(self):
        assert _escape_md_table("") == ""

    def test_none_returns_none(self):
        assert _escape_md_table(None) is None


# --- Tests _sanitize_url ---


class TestSanitizeUrl:
    def test_http_url_preserved(self):
        assert _sanitize_url("https://example.com") == "https://example.com"

    def test_http_url_preserved_no_scheme(self):
        assert _sanitize_url("example.com") == "example.com"

    def test_javascript_blocked(self):
        assert _sanitize_url("javascript:alert(1)") == "#"

    def test_data_blocked(self):
        assert _sanitize_url("data:text/html,<h1>test</h1>") == "#"

    def test_vbscript_blocked(self):
        assert _sanitize_url("vbscript:msgbox(1)") == "#"

    def test_file_blocked(self):
        assert _sanitize_url("file:///etc/passwd") == "#"

    def test_blob_blocked(self):
        assert _sanitize_url("blob:https://example.com/uuid") == "#"

    def test_empty_url_preserved(self):
        assert _sanitize_url("") == ""

    def test_none_url_preserved(self):
        assert _sanitize_url(None) is None


# --- Tests convert_xlsx_to_md ---


class TestConvertXlsxToMd:
    def test_simple_sheet_conversion(self, sample_xlsx_bytes):
        result = convert_xlsx_to_md(sample_xlsx_bytes)
        assert "# Feuille1" in result
        assert "| Nom | Valeur |" in result
        assert "| --- | --- |" in result
        assert "| A | 1 |" in result
        assert "| B | 2 |" in result

    def test_multi_sheet_conversion(self, sample_xlsx_multi_sheet_bytes):
        result = convert_xlsx_to_md(sample_xlsx_multi_sheet_bytes)
        assert "# Données" in result
        assert "# Résumé" in result
        assert "| Produit | Prix |" in result
        assert "| Pomme | 1.5 |" in result
        assert "| Orange | 2 |" in result  # openpyxl converts 2.0 to 2
        assert "| Total | 3.5 |" in result

    def test_merged_cells(self, sample_xlsx_merged_cells_bytes):
        result = convert_xlsx_to_md(sample_xlsx_merged_cells_bytes)
        assert "# Fusionné" in result
        assert "En-tête fusionné" in result
        assert "| Col1 | Col2 | Col3 |" in result

    def test_dates_and_numbers(self, sample_xlsx_dates_numbers_bytes):
        result = convert_xlsx_to_md(sample_xlsx_dates_numbers_bytes)
        assert "# Types" in result
        assert "42.5" in result
        assert "-10" in result
        assert "2024-01-15" in result
        assert "2024-06-30T12:00:00" in result
        assert "true" in result
        assert "false" in result

    def test_special_chars_escaped(self, sample_xlsx_special_chars_bytes):
        result = convert_xlsx_to_md(sample_xlsx_special_chars_bytes)
        assert "# Spécial" in result
        assert "\\|" in result  # pipe should be escaped
        assert "\\\\" in result  # backslash should be escaped

    def test_empty_sheet(self, sample_xlsx_empty_sheet_bytes):
        result = convert_xlsx_to_md(sample_xlsx_empty_sheet_bytes)
        assert "# Vide" in result
        # Empty sheets just show an empty table structure
        assert (
            "|  |" in result
            or "_empty_" in result.lower()
            or "(empty)" in result.lower()
        )

    def test_empty_cells(self):
        """Test that empty cells are handled correctly."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        ws["A1"] = "Header"
        ws["B1"] = "Data"
        ws["A2"] = "value"
        # ws["B2"] is empty
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = convert_xlsx_to_md(buf.getvalue())
        assert "| value |  |" in result  # Empty cell has a space before the closing |

    def test_invalid_xlsx_raises_error(self):
        """Test that invalid XLSX raises an error."""
        with pytest.raises(Exception, match=".*"):
            convert_xlsx_to_md(b"not a valid xlsx file")

    def test_boolean_values(self):
        """Test boolean cell values."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bool"
        ws["A1"] = "Col"
        ws["B1"] = "Bool"
        ws["A2"] = "test"
        ws["B2"] = True
        ws["A3"] = "test2"
        ws["B3"] = False
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = convert_xlsx_to_md(buf.getvalue())
        assert "true" in result
        assert "false" in result


# --- Integration tests via API ---


class TestXlsxApiIntegration:
    @pytest.mark.asyncio
    async def test_upload_valid_xlsx(self, async_client, sample_xlsx_bytes):
        """Upload valid XLSX → 200 OK with Markdown."""
        response = await async_client.post(
            "/convert",
            files={
                "file": (
                    "test.xlsx",
                    sample_xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert "markdown" in data
        assert "# Feuille1" in data["markdown"]

    @pytest.mark.asyncio
    async def test_upload_invalid_xlsx_wrong_format(self, async_client):
        """Upload XLSX with wrong format → 400."""
        response = await async_client.post(
            "/convert",
            files={
                "file": ("test.xlsx", b"not an xlsx file", "application/octet-stream")
            },
        )
        assert response.status_code == 400

    @pytest.mark.asyncio
    async def test_upload_xlsx_corrupted(self, async_client):
        """Upload corrupted XLSX → 400."""
        response = await async_client.post(
            "/convert",
            files={"file": ("test.xlsx", b"PK\x03\x04corrupted", "application/zip")},
        )
        assert response.status_code == 400

    @pytest.mark.asyncio
    async def test_upload_xlsx_empty(self, async_client):
        """Upload empty XLSX → 400."""
        response = await async_client.post(
            "/convert", files={"file": ("test.xlsx", b"", "application/octet-stream")}
        )
        assert response.status_code == 400

    @pytest.mark.asyncio
    async def test_upload_xlsx_multi_sheet(
        self, async_client, sample_xlsx_multi_sheet_bytes
    ):
        """Upload multi-sheet XLSX → Markdown with headings."""
        response = await async_client.post(
            "/convert",
            files={
                "file": (
                    "test.xlsx",
                    sample_xlsx_multi_sheet_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert "# Données" in data["markdown"]
        assert "# Résumé" in data["markdown"]
